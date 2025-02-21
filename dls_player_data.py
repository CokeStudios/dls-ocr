from string import ascii_uppercase, ascii_letters
from time import perf_counter
from tkinter import Tk, StringVar, Label, Entry, Button
import os

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from PIL import Image, ImageTk
from tqdm import tqdm
import easyocr
import numpy

READER = easyocr.Reader(['en', 'ch_sim'])
DEVICE = {
    '1536x2048': {
        'device': ('IPAD'),
        'cards_count': (3, 3),
        'topleft_coords': (77, 284),
        'card_width': 423,
        'width_space': 27,
        'card_height': 260,
        'height_space': 43,
        'name': (0, 0, 423, 41),
        # 'stats': (80, 16),
        'stats_topleft_coords': (8, 71),
        'stats_size': 53,
        'stats_width_space': 6,
        'stats_height_space': 33,
        'overall': (275, 64, 306, 84),
        'height': (10, 234, 42, 256),
        'leg': (83, 234, 105, 256),
        'price': (300, 222, 420, 260),
        'position': (330, 172, 352, 190)
    },
    '828x1792': {
        'device': ('iPhone XR'),
        'cards_count': (3, 3),
        'topleft_coords': (186, 172),
        'card_width': 257,
        'width_space': 102,
        'card_height': 159,
        'height_space': 25,
        'name': (0, 0, 257, 26),
        'stats_topleft_coords': (6, 43),
        'stats_size': 33,
        'stats_width_space': 2,
        'stats_height_space': 19,
        'overall': (163, 33, 189, 59),
        'height': (8, 143, 26, 155),
        'leg': (52, 143, 86, 157),
        'price': (159, 134, 229, 158),
        'position': (200, 106, 215, 115)
    }
}


def check_overall(pixel: tuple[int]):
    check1 = 38 <= pixel[0] <= 41 and 220 <= pixel[1] <= 222 and \
        20 <= pixel[2] <= 30
    check2 = 240 <= pixel[0] <= 243 and 197 <= pixel[1] <= 200 and \
        76 <= pixel[2] <= 78
    check3 = 240 <= pixel[0] <= 255 and 135 <= pixel[1] <= 140 and \
        20 <= pixel[2] <= 26
    return check1 or check2 or check3


def check_position(pixel: tuple[int]):
    check1 = pixel[0] == 255 and 85 <= pixel[1] <= 90 and \
        115 <= pixel[2] <= 125
    check2 = 45 <= pixel[0] <= 55 and 165 <= pixel[1] <= 175 and \
        250 <= pixel[2] <= 255
    check3 = 250 <= pixel[0] <= 255 and 250 <= pixel[1] <= 255 and \
        85 <= pixel[2] <= 95
    check4 = 65 <= pixel[0] <= 75 and 245 <= pixel[1] <= 255 and \
        90 <= pixel[2] <= 100
    return check1 or check2 or check3 or check4


def scan_coords(image: Image.Image) -> dict:
    device_dict = {'cards_count': (3, 3)}
    width, height = image.size
    for x in range(width):
        pixels = []
        have = False
        have2 = False
        for y in range(height):
            pixel = image.getpixel((x, y))[0:3]
            if pixel == (133, 133, 133):
                have = True
                pixels.append(pixel)
                if len(pixels) >= 20:
                    have2 = True
            elif len(pixels) >= 3 and pixel == (98, 98, 98) and \
                    'topleft_coords' not in device_dict:
                device_dict['topleft_coords'] = [x]
            else:
                pixels = []

        if have is False and 'topleft_coords' in device_dict and \
                'card_width' not in device_dict:
            device_dict['card_width'] = x - 1 - \
                device_dict['topleft_coords'][0]

        if have2 is True and 'card_width' in device_dict:
            device_dict['width_space'] = x - 1 - \
                device_dict['topleft_coords'][0] - device_dict['card_width']
            break

    else:
        return {}

    for y in range(height):
        pixels = []
        pixels2 = []
        for x in range(width):
            pixel = image.getpixel((x, y))[0:3]
            if all(132 <= x <= 134 for x in pixel):
                pixels.append(pixel)
            else:
                pixels = []

            if len(pixels) == 200 and len(device_dict['topleft_coords']) == 1:
                device_dict['topleft_coords'].append(y)
                for y1 in range(y + 1, y + 101):
                    pixel1 = image.getpixel((x, y1))[0:3]
                    if all(72 <= x <= 98 for x in pixel1):
                        device_dict['name'] = \
                            (0, 0, device_dict['card_width'],
                                y1 - device_dict['topleft_coords'][1])
                        break
                break

            if all(53 <= x <= 57 for x in pixel):
                pixels2.append(pixel)
            else:
                pixels2 = []

            if len(pixels2) == 100 and 'height_space' not in device_dict:
                for y2 in range(y + 11, y + 81):
                    pixel3 = image.getpixel((x, y2))[0:3]
                    if 'panel' not in device_dict and \
                            all(53 <= x <= 57 for x in pixel3):
                        device_dict['panel'] = y2 - 1 - \
                            device_dict['topleft_coords'][1]

                    if all(18 <= x <= 30 for x in pixel3) and \
                            'card_height' not in device_dict and \
                            len(device_dict['topleft_coords']) == 2:
                        device_dict['card_height'] = \
                            y2 - device_dict['topleft_coords'][1]
                        break

                for y3 in range(y2 + 21, y2 + 200):
                    pixel4 = image.getpixel((x, y3))[0:3]
                    if all(123 <= x <= 135 for x in pixel4) and \
                            'card_height' in device_dict:
                        device_dict['height_space'] = y3 - y2
                        break

    x0 = device_dict['topleft_coords'][0]
    y0 = device_dict['topleft_coords'][1]
    y0 += device_dict['card_height'] + device_dict['height_space']
    x1 = x0 + device_dict['card_width']
    y1 = y0 + device_dict['card_height']
    image = image.crop((x0, y0, x1, y1))
    width, height = image.size
    sv = 0
    for x in range(2, width - 2):
        have = False
        for y in range(4, height - 4):
            pixel = image.getpixel((x, y))[0:3]
            if pixel == (0, 0, 0):
                have = True
                break

        if have is True and \
                'stats_topleft_coords' not in device_dict:
            device_dict['stats_topleft_coords'] = [x]

        if have is False and 'stats_topleft_coords' in device_dict and \
                'stats_size' not in device_dict:
            sv = x
            device_dict['stats_size'] = x - 1 - \
                device_dict['stats_topleft_coords'][0]

        if have is True and 'stats_size' in device_dict:
            device_dict['stats_width_space'] = x - sv
            break

    sv2 = 0
    for y in range(4, height - 4):
        have = False
        for x in range(2, width - 2):
            pixel = image.getpixel((x, y))[0:3]
            if pixel == (0, 0, 0):
                have = True
                break

        if have is True and \
                len(device_dict['stats_topleft_coords']) == 1:
            device_dict['stats_topleft_coords'].append(y)

        if have is False and len(device_dict['stats_topleft_coords']) == 2 \
                and sv2 == 0:
            sv2 = y

        if have is True and len(device_dict['stats_topleft_coords']) == 2 \
                and sv2 != 0:
            device_dict['stats_height_space'] = y - sv2
            break

    last = (0, 0, 0)
    for y in range(height):
        if 'overall' in device_dict and 'position' in device_dict:
            break

        for x in range(width):
            pixel = image.getpixel((x, y))[0:3]
            if check_overall(pixel) and 'overall' not in device_dict:
                device_dict['overall'] = [
                    0, y, 0, int(y + device_dict['stats_size'] * 0.7) + 1]
                break

            if check_position(pixel) and 'position' not in device_dict:
                device_dict['position'] = [x, y, 0, 0]

            if check_position(last) and (not check_position(pixel)) and \
                    device_dict['position'][2] == 0:
                device_dict['position'][2] = x
                break

            last = pixel[::]

    last = (0, 0, 0)
    for x in range(width):
        if device_dict['overall'][0] != 0 and device_dict['position'][3] != 0:
            break

        for y in range(height):
            pixel = image.getpixel((x, y))[0:3]
            if check_overall(pixel) and device_dict['overall'][0] == 0:
                device_dict['overall'][0] = x
                device_dict['overall'][2] = x + \
                    int(device_dict['stats_size'] * .7) + 1
                break

            if check_position(last) and (not check_position(pixel)) and \
                    device_dict['position'][3] == 0:
                device_dict['position'][3] = y
                break

            last = pixel[::]

    device_dict['panel'] = (0, device_dict['panel'],
                            image.width, image.height)
    return device_dict


def parse_image(image_dir: str,
                max_file: int = -1, max_cards: int = -1,
                output: bool = False, rename: bool = True,
                restore: bool = False):
    '''
    Input: Transfer market screenshot directory
    Output: Players and stats
    '''
    if restore is True:
        for image_file in os.listdir(image_dir):
            filename = image_dir + os.sep + image_file
            fn2 = filename.replace('_OLD', '')
            os.rename(filename, fn2)

    cards: list[Image.Image] = []
    i = 0
    for image_file in os.listdir(image_dir):
        filename = image_dir + os.sep + image_file
        if '.' not in filename or image_file[0] == '.':
            continue

        if filename.split('.')[-2].endswith('_OLD'):
            continue

        if i == max_file:
            break

        i += 1

        try:
            image = Image.open(filename)
            if rename is not False:
                fn = filename.split('.')
                os.rename(filename, fn[-2] + '_OLD.' + fn[-1])
        except Exception:
            continue

        device_dict = scan_coords(image)
        x0 = device_dict['topleft_coords'][0]
        count_x = 0
        while count_x < device_dict['cards_count'][0]:
            y0 = device_dict['topleft_coords'][1]
            x1 = x0 + device_dict['card_width']
            count_y = 0
            while count_y < device_dict['cards_count'][1]:
                y1 = y0 + device_dict['card_height']
                card = image.crop((x0, y0, x1, y1))
                cards.append(card)
                count_y += 1
                y0 = y1 + device_dict['height_space']

            count_x += 1
            x0 = x1 + device_dict['width_space']

    if max_cards != -1:
        cards = cards[:max_cards]

    if output is not True:
        pbar = tqdm(cards)
    else:
        pbar = cards

    result: list[str] = []
    for card in pbar:
        t1 = perf_counter()
        name_image_ori = card.crop(device_dict['name'])
        name_image = name_image_ori.resize(
            (name_image_ori.size[0] * 2, name_image_ori.size[1] * 2),
            resample=Image.BILINEAR)
        player_name: list[str] = READER.readtext(numpy.asarray(name_image),
                                                 detail=False)
        if player_name == [] or not bool(player_name[0].strip()) or \
                player_name[0].strip().lower() in ['神秘球员', 'secret player']:
            continue

        player_name = player_name[0].title().strip('_')
        k = check_has_player(player_name)
        if k is not None and k[1] is True:
            continue

        if output is not True:
            pbar.set_description(f'Processing {player_name}')

        stats = []
        stats_images = []
        x0 = device_dict['stats_topleft_coords'][0]
        count_x = 0
        while count_x < 4:
            y0 = device_dict['stats_topleft_coords'][1]
            x1 = x0 + device_dict['stats_size']
            count_y = 0
            while count_y < 2:
                y1 = y0 + device_dict['stats_size']
                stat = card.crop((x0, y0, x1, y1))
                stats_images.append(stat)
                # pixels = stat.load()
                # width, height = stat.size
                # for y in range(height):
                #     for x in range(width):
                #         if 150 <= sum(pixels[x, y]) <= 210:
                #             pixels[x, y] = (255, 255, 255)
                #         else:
                #             pixels[x, y] = (0, 0, 0)

                stat_text = READER.readtext(numpy.asarray(stat), detail=False)
                if stat_text in [[], ['']]:
                    pixels = stat.load()
                    width, height = stat.size
                    for y in range(height):
                        for x in range(width):
                            if pixels[x, y][0] >= 150 and \
                                    sum(pixels[x, y][1:3]) <= 100:
                                pixels[x, y] = (255, 255, 255)
                            else:
                                pixels[x, y] = (0, 0, 0)

                    stat = stat.resize((stat.size[0] * 4, stat.size[1] * 4),
                                       resample=Image.BILINEAR)
                    stat_text = READER.readtext(numpy.asarray(stat),
                                                detail=False)
                    if stat_text in [[], ['']]:
                        stat_text = ['']
                        stat = stat.resize(
                            (stat.size[0] * 6, stat.size[1] * 6),
                            resample=Image.NEAREST)
                        stat_text = READER.readtext(numpy.asarray(stat),
                                                    detail=False)
                        if stat_text in [[], ['']]:
                            stat_text = ['']

                stats.append(stat_text[0])
                count_y += 1
                y0 = y1 + device_dict['stats_height_space']

            count_x += 1
            x0 = x1 + device_dict['stats_width_space']

        temp_stats = [stats[0], stats[2], stats[4], stats[6],
                      stats[1], stats[3], stats[5], stats[7]]
        stats_images = [stats_images[0], stats_images[2],
                        stats_images[4], stats_images[6],
                        stats_images[1], stats_images[3],
                        stats_images[5], stats_images[7]]
        stats = []
        for x in temp_stats:
            try:
                stats.append(int(x))
            except ValueError:
                stats.append(0)

        player_name = [x.replace('。', '.').replace('Vinijr.',
                                                   'Vinicius Junior')
                       for x in player_name.strip().split(' ', maxsplit=1)]
        club, nationality = '', ''
        overall_image = card.crop(device_dict['overall'])
        overall_ori = overall_image.copy()
        overall_image = overall_image.resize(
            (overall_image.size[0] * 6, overall_image.size[1] * 6),
            resample=Image.NEAREST)
        overall = READER.readtext(numpy.asarray(overall_image),
                                  detail=False)
        if overall in [[], ['']]:
            pixels = overall_image.load()
            width, height = overall_image.size
            for y in range(height):
                for x in range(width):
                    if sum(pixels[x, y][0:3]) >= 600:
                        pixels[x, y] = (0, 0, 0)
                    else:
                        pixels[x, y] = (255, 255, 255)

            overall = READER.readtext(numpy.asarray(overall_image),
                                      detail=False)

        overall = overall[0].strip()
        try:
            panel_image = card.crop(device_dict['panel'])
            txt = READER.readtext(numpy.asarray(panel_image),
                                  detail=False)
            height = txt[0].strip().strip(ascii_letters)
            leg = txt[1].strip().lower()
            price = txt[2].strip()
            # height_image = card.crop(device_dict['height'])
            # height = READER.readtext(numpy.asarray(height_image),
            #                          detail=False)[0].strip()
            # leg_image = card.crop(device_dict['leg'])
            # leg = READER.readtext(numpy.asarray(leg_image),
            #                       detail=False)[0].strip().lower()
            # price_image = card.crop(device_dict['price'])
            # price: str = READER.readtext(numpy.asarray(price_image),
            #                              detail=False)[0].strip()
            price = price.replace(',', '').replace('，', '').replace(' ', '')
            if leg in ['left', '左']:
                leg = 'L'
            elif leg in ['right', '右']:
                leg = 'R'
            elif leg in ['both', '双', '双脚']:
                leg = 'B'

            position_image = card.crop(device_dict['position'])
            pos_image2 = position_image.resize(
                (position_image.size[0] * 4, position_image.size[1] * 4),
                resample=Image.BILINEAR)
            position: str = READER.readtext(numpy.asarray(pos_image2),
                                            detail=False)
            if position in [[], ['']]:
                position: str = READER.readtext(numpy.asarray(position_image),
                                                detail=False)
                if position in [[], ['']]:
                    position = ['']

            position = position[0].strip().upper()
            pos_table = str.maketrans({'N': 'M', '8': 'B',
                                       '1': 'L', 'I': '', '[': 'L'})
            position = position.translate(pos_table)

            t2 = perf_counter()
            panel_image_1 = panel_image.crop(
                (0, 0, panel_image.width // 2, panel_image.height))
            panel_image_2 = panel_image.crop(
                (panel_image.width // 2, 0,
                 panel_image.width, panel_image.height))
            player_tuple = ((card, name_image_ori, overall_ori,
                             position_image, stats_images,
                             panel_image_1, panel_image_1, panel_image_2),
                            (club, nationality), ' '.join(player_name),
                            int(overall), position, stats,
                            int(height), leg, int(price),
                            f'Time: {round(t2 - t1, 2)}s')
            result.append(player_tuple)
            if output is True:
                print(player_tuple[1:][1:])

        except Exception:
            card.show()

    return result


# def on_key(event):
#     keycode = event.keycode
#     if keycode in [13, 32]:
#         root.quit()


def get_font(number):
    if number == 'NEW':
        color = '00ff00'
    else:
        number = int(number)
        if number >= 90:
            color = '00ffff'
        elif number >= 80:
            color = '00ff00'
        elif number >= 70:
            color = 'ffff00'
        elif number >= 60:
            color = 'ff9a00'
        else:
            color = 'ff0000'

    font = Font(name='Arial', size=11, bold=True, color=color)
    return font


def get_font_2(number):
    if number == 'NEW':
        color = '00ff00'
    else:
        number = int(number)
        if number > 0:
            color = '00ff00'
        elif number < 0:
            color = 'ff0000'
        else:
            color = 'ffffff'

    font = Font(name='Arial', size=11, bold=True, color=color)
    return font


def get_fill(position):
    if position == 'GK':
        color = '6d9dca'
    elif position in ['CF', 'LW', 'RW', 'SS']:
        color = 'd65452'
    elif position in ['LB', 'CB', 'RB']:
        color = '49b147'
    else:
        color = 'f3d15e'

    fill = PatternFill('solid', fgColor=color)
    return fill


def check_has_player(player_name):
    lst = [wb['Legendary Players'], wb['Rare Players'], wb['Common Players']]
    for ws1 in lst:
        players = ws1['B':'C']
        player_names = []
        for x in zip([x.value for x in players[0][3:]],
                     [x.value for x in players[1][3:]]):
            if x == (None, None):
                continue
            if x[0] is None:
                x = ('', x[1])
            player_names.append(' '.join(list(x)).lower())

        try:
            index = player_names.index(player_name.lower()) + 4
            fill = ws1[f'A{index}'].fill
            updated = False
            if fill.fgColor.rgb[2:].lower() == '00ff00':
                updated = True
            return (index, ws1), updated

        except ValueError:
            pass


def write_player_data(data):
    for data_list in data:
        k = check_has_player(data_list[0])
        if k is not None:
            index, ws1 = k[0]
            player_id = ws1[f'W{index}'].value
            old_rating = int(ws1[f'I{index}'].value)
            old_nat = ws1[f'E{index}'].value
            old_club = ws1[f'F{index}'].value
            ws1.delete_rows(index)
        else:
            player_id = old_nat = old_club = ''
            old_rating = 0

        ws = wb['Legendary Players']
        ws.sheet_properties.tabColor = 'f5bd00'
        ws = wb['Rare Players']
        ws.sheet_properties.tabColor = '349bf9'
        ws = wb['Common Players']
        ws.sheet_properties.tabColor = 'cccccc'

        rating = int(data_list[1])
        if rating >= 80:
            ws = wb['Legendary Players']
        elif rating >= 70:
            ws = wb['Rare Players']
        else:
            ws = wb['Common Players']

        stats = list(map(int, data_list[3:11]))
        # total_stats = sum(stats)
        stats += ['', '']
        if data_list[2] == 'GK':
            stats[8] = stats[2]
            stats[2] = ''
            stats[9] = stats[6]
            stats[6] = ''

        if old_rating == 0:
            rating_change = 'NEW'
        else:
            rating_change = rating - old_rating

        player_name_list = data_list[0].split(' ', maxsplit=1)
        if len(player_name_list) == 1:
            player_name = ['', player_name_list[0]]
        else:
            player_name = [player_name_list[0], player_name_list[1]]

        row = ws.max_row + 1
        input_data = ['', *player_name, data_list[-3],
                    #   data_list[-2], data_list[-1], data_list[2],
                      old_nat, old_club, data_list[2],
                      data_list[-4], rating, data_list[-5], *stats,
                      rating_change,  # total_stats,
                      '', player_id]
        ws.append(input_data)

        ws[f'A{row}'].fill = PatternFill('solid', fgColor='00ff00')
        font1 = Font(name='Arial', size=11, bold=True, color='ffffff')
        font2 = Font(name='Arial', size=11, bold=True, color='000000')
        align1 = Alignment(horizontal='center', vertical='bottom')
        align2 = Alignment(horizontal='general', vertical='bottom')
        for x in ws[f'B{row}:W{row}']:
            for y in x:
                y.fill = PatternFill('solid', fgColor='000000')
        for x in ws[f'B{row}:F{row}']:
            for y in x:
                y.font = font1
        for x in ws[f'B{row}:D{row}']:
            for y in x:
                y.alignment = align1
        for x in ws[f'E{row}:F{row}']:
            for y in x:
                y.alignment = align2
        ws[f'G{row}'].font = font2
        ws[f'G{row}'].fill = get_fill(data_list[2])
        for x in ws[f'G{row}:W{row}']:
            for y in x:
                y.alignment = align1
        ws[f'H{row}'].font = font1
        ws[f'J{row}'].font = font1
        for x in ws[f'V{row}:W{row}']:
            for y in x:
                y.font = font1

        for i in range(8, 20):
            if i == 9:
                continue
            number = input_data[i]
            if number == '':
                continue

            font = get_font(number)
            cell = f'{ascii_uppercase[i]}{row}'
            ws[cell].font = font

        font = get_font_2(input_data[20])
        cell = f'U{row}'
        ws[cell].font = font

    for ws in [wb['Legendary Players'], wb['Rare Players'],
               wb['Common Players']]:
        for i in range(3, len(ws['V'])):
            i += 1
            ws[f'V{i}'] = f'=SUM(K{i}:T{i})'

    wb.save(empty_database)
    wb.close()


all_data = []


def callback():
    r = []
    for x in var:
        r.append(x.get())

    all_data.append(r)
    root.destroy()


def check_gui(data: list[list[list[Image.Image]]]):
    '''
    A gui to check the data.
    '''
    global root, var
    for stats_list in data:
        root = Tk()
        root.title('Data Checker')
        root.resizable(width=False, height=False)

        var = []
        for x in range(16):
            v = StringVar(root)
            var.append(v)

        # card_image = ImageTk.PhotoImage(stats_list[0][0])
        # panel = Label(root, image=card_image)
        # panel.grid(row=1, column=1)

        image_list = stats_list[0][1:]
        image_list_ = image_list[::]
        image_list = []
        for a in image_list_:
            if isinstance(a, (list, tuple)):
                for b in a:
                    image_list.append(b)
            else:
                image_list.append(a)

        stats_list_ = stats_list[2:]
        new_stats_list = []
        for a in stats_list_:
            if isinstance(a, (list, tuple)):
                for b in a:
                    new_stats_list.append(b)
            else:
                new_stats_list.append(a)

        rows = 0
        row = 0
        column_minus = 0
        for i, image in enumerate(image_list):
            if rows > 0 and rows % 4 == 0:
                rows = 0
                row += 2
                column_minus = i

            image = ImageTk.PhotoImage(image)
            panel = Label(root, image=image)
            panel.image = image
            panel.grid(row=row, column=i - column_minus, sticky='w')

            if i in [0, 1, 2, 11]:
                text = Entry(root, textvariable=var[i], bg='orange')
            elif new_stats_list[i] in [0, '0', '']:
                text = Entry(root, textvariable=var[i], bg='red')
            else:
                text = Entry(root, textvariable=var[i])

            text.insert(0, new_stats_list[i])
            text.grid(row=row + 1, column=i - column_minus, sticky='w')
            rows += 1

        button = Button(root, text='Submit', default='active',
                        command=callback)
        button.grid(row=row + 2, column=0, sticky='w')
        button2 = Button(root, text='Skip', command=root.destroy)
        button2.grid(row=row + 2, column=1, sticky='w')
        root.mainloop()

    write_player_data(all_data)


if __name__ == '__main__':
    # Will rename all the image files (so that it won't be parsed next time)
    # param "restore=True" will reset images' filenames
    empty_database = 'DLS 25 test database.xlsx'
    image_dir = 'dls25/winter'

    wb = load_workbook(empty_database)
    result = parse_image(image_dir, rename=False,
                         output=True)
    check_gui(result)
